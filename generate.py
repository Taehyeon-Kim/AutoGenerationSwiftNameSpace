import os
import sys
import urllib.request
import importlib
import csv
import openpyxl


importlib.reload(sys)

gdoc_id = "1WMnHk1brX7CR34hCN88cDnTDc4UgTAKh-qegwnPIqCU/edit#gid=0"


def get_gdoc_information():
    download_path = sys.argv[1]
    try:
        csv_file = export_xlsx_from_sheet(gdoc_id)
        for sheet_name in get_multiple_sheet(csv_file):
            string_list = multiple_sheet(csv_file, sheet_name)
            write_strings(sheet_name, string_list, download_path)
        os.remove(csv_file)
    except Exception as e:
        print(":::::::::::::ERROR:::::::::::::")
        print(e)


def export_xlsx_from_sheet(gdoc_id, download_path=None, ):
    print("Downloading the XLSX file with id %s" % gdoc_id)

    resource = gdoc_id.split('/')[0]
    resource_id = 'spreadsheet:' + resource

    if download_path is None:
        download_path = os.path.abspath(os.path.dirname(__file__))

    file_name = os.path.join(download_path, '%s.xlsx' % (resource))

    print('download_path : %s' % download_path)
    print('Downloading spreadsheet to %s' % file_name)

    url = 'https://docs.google.com/spreadsheet/ccc?key=%s&output=xlsx' % (
        resource)
    urllib.request.urlretrieve(url, file_name)

    print("Download Completed!")

    return file_name
    
def get_multiple_sheet(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws_names = wb.sheetnames
    print(ws_names)
    return ws_names

def multiple_sheet(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]
    wr = sheet.rows
    next(wr)
    cols = check_row_column(sheet)
    string_list = []
    
    for row in wr:
        key = row[cols["key"]-1].value
        kr = row[cols["kr"]-1].value
        dict_string = {
            "key": key,
            "kr": kr
        }
        string_list.append(dict_string)
    
    return string_list

    
        
def check_row_column(sheet):
    for col in range(1, sheet.max_column + 1):
        for row in range(1, sheet.max_row + 1):
            data = sheet.cell(row = row, column = col).value
            if data == "key":
                key = col
            elif data == "kr":
                kr = col
            
    return {"key" : key, "kr" : kr}

def write_strings(filename, string_list, save_path):
    swift_file = open(filename+".swift", "w")

    swift_file.write("import UIKit\n\n")
    swift_file.write("enum "+ filename + " {\n")

    for item in string_list:
        swift_file.write("\tstatic let " + item["key"] + " = " + "\"" + item["kr"] + "\"\n")

    swift_file.write("}")
    swift_file.close()

if __name__ == '__main__':
    get_gdoc_information()
